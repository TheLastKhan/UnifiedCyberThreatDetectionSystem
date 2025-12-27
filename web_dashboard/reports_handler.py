"""
Reports Handler - Excel/PDF Export/Import functionality
"""

import io
import json
from datetime import datetime
from typing import List, Dict, Any, Tuple
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib import colors


def export_predictions_to_excel(email_predictions: List[Dict[str, Any]], web_predictions: List[Dict[str, Any]]) -> io.BytesIO:
    """Export predictions to Excel file"""
    wb = Workbook()
    
    # Email Predictions Sheet
    ws_email: Worksheet = wb.active  # type: ignore
    if ws_email is not None:
        ws_email.title = "Email Predictions"
    
    if ws_email is not None:
        ws_email.title = "Email Predictions"
        
        # Headers
        headers = ['ID', 'Timestamp', 'Prediction', 'Confidence', 'Risk Level', 'Subject', 'Sender']
        ws_email.append(headers)
        
        # Style headers
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col_num, cell in enumerate(ws_email[1], 1):
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Add data
        for pred in email_predictions:
            ws_email.append([
                pred.get('id', ''),
                pred.get('timestamp', ''),
                pred.get('prediction', ''),
                pred.get('confidence', 0),
                pred.get('risk_level', ''),
                pred.get('email_subject', '')[:50] if pred.get('email_subject') else '',
                pred.get('email_sender', '')
            ])
        
        # Auto-adjust column widths
        for column in ws_email.columns:
            max_length = 0
            try:
                column_letter = column[0].column_letter
            except (AttributeError, IndexError):
                column_letter = 'A'
            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_email.column_dimensions[column_letter].width = adjusted_width
    
    # Web Predictions Sheet
    ws_web: Worksheet = wb.create_sheet(title="Web Predictions")  # type: ignore
    
    headers = ['ID', 'Timestamp', 'Is Anomaly', 'Anomaly Score', 'IP Address', 'Patterns Detected']
    ws_web.append(headers)
    
    # Style headers
    for col_num, cell in enumerate(ws_web[1], 1):
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Add data
    for pred in web_predictions:
        patterns = ', '.join(pred.get('patterns_detected', [])) if pred.get('patterns_detected') else ''
        ws_web.append([
            pred.get('id', ''),
            pred.get('timestamp', ''),
            'Yes' if pred.get('is_anomaly') else 'No',
            pred.get('anomaly_score', 0),
            pred.get('ip_address', ''),
            patterns[:100]
        ])
    
    # Auto-adjust column widths
    for column in ws_web.columns:
        max_length = 0
        try:
            column_letter = column[0].column_letter
        except (AttributeError, IndexError):
            column_letter = 'A'
        for cell in column:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_web.column_dimensions[column_letter].width = adjusted_width
    
    # Summary Sheet
    ws_summary = wb.create_sheet(title="Summary")
    ws_summary.append(['Threat Detection Report'])
    ws_summary.append(['Generated:', datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
    ws_summary.append([])
    ws_summary.append(['Total Email Predictions:', len(email_predictions)])
    ws_summary.append(['Total Web Predictions:', len(web_predictions)])
    ws_summary.append(['Total Threats:', len(email_predictions) + len(web_predictions)])
    
    # Style summary
    ws_summary['A1'].font = Font(size=16, bold=True)
    ws_summary.column_dimensions['A'].width = 30
    ws_summary.column_dimensions['B'].width = 20
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output


def import_predictions_from_excel(file_stream: Any) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    """Import predictions from Excel file"""
    wb = load_workbook(file_stream)
    
    email_predictions: List[Dict[str, Any]] = []
    web_predictions: List[Dict[str, Any]] = []
    
    # Read Email Predictions
    if 'Email Predictions' in wb.sheetnames:
        ws = wb['Email Predictions']
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        
        for row in rows:
            if row[0]:  # If ID exists
                email_predictions.append({
                    'prediction': str(row[2]) if row[2] else '',
                    'confidence': float(row[3]) if row[3] and isinstance(row[3], (int, float)) else 0.5,
                    'risk_level': str(row[4]) if row[4] else '',
                    'email_subject': str(row[5]) if row[5] else '',
                    'email_sender': str(row[6]) if row[6] else ''
                })
    
    # Read Web Predictions
    if 'Web Predictions' in wb.sheetnames:
        ws = wb['Web Predictions']
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        
        for row in rows:
            if row[0]:  # If ID exists
                patterns_str = str(row[5]) if row[5] else ''
                patterns = patterns_str.split(', ') if patterns_str else []
                web_predictions.append({
                    'is_anomaly': row[2] == 'Yes',
                    'anomaly_score': float(row[3]) if row[3] and isinstance(row[3], (int, float)) else 0.5,
                    'ip_address': str(row[4]) if row[4] else '',
                    'patterns_detected': patterns
                })
    
    return email_predictions, web_predictions


def export_predictions_to_pdf(email_predictions: List[Dict[str, Any]], web_predictions: List[Dict[str, Any]]) -> io.BytesIO:
    """Export predictions to PDF file"""
    buffer = io.BytesIO()
    
    doc = SimpleDocTemplate(buffer, pagesize=letter,
                           rightMargin=72, leftMargin=72,
                           topMargin=72, bottomMargin=18)
    
    story = []
    styles = getSampleStyleSheet()
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#1f2937'),
        spaceAfter=30,
        alignment=1  # Center
    )
    
    story.append(Paragraph("Threat Detection Report", title_style))
    story.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles['Normal']))
    story.append(Spacer(1, 0.3*inch))
    
    # Summary
    summary_data = [
        ['Total Email Predictions:', str(len(email_predictions))],
        ['Total Web Predictions:', str(len(web_predictions))],
        ['Total Threats:', str(len(email_predictions) + len(web_predictions))]
    ]
    
    summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#f3f4f6')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 12),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey)
    ]))
    
    story.append(summary_table)
    story.append(Spacer(1, 0.5*inch))
    
    # Email Predictions Section
    story.append(Paragraph("Email Predictions", styles['Heading2']))
    story.append(Spacer(1, 0.2*inch))
    
    if email_predictions:
        email_data = [['Prediction', 'Confidence', 'Risk', 'Subject']]
        for pred in email_predictions[:20]:  # Limit to 20
            email_data.append([
                pred.get('prediction', 'N/A'),
                f"{pred.get('confidence', 0)*100:.1f}%",
                pred.get('risk_level', 'N/A'),
                pred.get('email_subject', '')[:30] + '...' if len(pred.get('email_subject', '')) > 30 else pred.get('email_subject', 'N/A')
            ])
        
        email_table = Table(email_data, colWidths=[1.5*inch, 1*inch, 1*inch, 3*inch])
        email_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey)
        ]))
        
        story.append(email_table)
    else:
        story.append(Paragraph("No email predictions found.", styles['Normal']))
    
    story.append(Spacer(1, 0.5*inch))
    
    # Web Predictions Section
    story.append(Paragraph("Web Predictions", styles['Heading2']))
    story.append(Spacer(1, 0.2*inch))
    
    if web_predictions:
        web_data = [['Anomaly', 'Score', 'IP Address', 'Patterns']]
        for pred in web_predictions[:20]:  # Limit to 20
            patterns_value = pred.get('patterns_detected')
            
            # Handle different types of patterns_detected values
            if not patterns_value or patterns_value == 'None':
                patterns = 'No patterns detected'
            elif isinstance(patterns_value, list):
                patterns = ', '.join(str(p) for p in patterns_value)[:40] if patterns_value else 'No patterns detected'
            else:
                patterns = str(patterns_value)[:40] if patterns_value else 'No patterns detected'
            
            web_data.append([
                'Yes' if pred.get('is_anomaly') else 'No',
                f"{pred.get('anomaly_score', 0):.2f}",
                pred.get('ip_address', 'N/A'),
                patterns
            ])
        
        web_table = Table(web_data, colWidths=[1*inch, 1*inch, 1.5*inch, 3*inch])
        web_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey)
        ]))
        
        story.append(web_table)
    else:
        story.append(Paragraph("No web predictions found.", styles['Normal']))
    
    # Build PDF
    doc.build(story)
    buffer.seek(0)
    
    return buffer
